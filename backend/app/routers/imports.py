import io
from typing import Literal

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse

from app import excel_store as store
from app.auth import get_current_user
from app.schemas import ImportBatch, ImportResult, ImportRowError

router = APIRouter(
    prefix="/import", tags=["import"], dependencies=[Depends(get_current_user)]
)

TEMPLATE_HEADERS = [
    "Name",
    "Designation",
    "Email",
    "Phone",
    "WhatsApp Number",
    "Place of Work",
    "Country",
    "Medical License No.",
    "Participant Type",
]

COLUMN_MAP = {
    "name": "name",
    "designation": "designation",
    "email": "email",
    "phone": "phone",
    "whatsapp number": "whatsapp_number",
    "place of work": "place_of_work",
    "country": "country",
    "medical license no.": "medical_license_no",
    "medical license no": "medical_license_no",
    "participant type": "participant_type",
}

REQUIRED_FIELDS = (
    "name",
    "designation",
    "email",
    "phone",
    "whatsapp_number",
    "place_of_work",
)


@router.get("/template")
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participants"
    ws.append(TEMPLATE_HEADERS)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cme_import_template.xlsx"},
    )


@router.post("/participants", response_model=ImportResult, status_code=201)
async def import_participants(
    event_id: str = Form(...),
    source_type: Literal["cme_website", "external_society"] = Form(...),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(
            status_code=400, detail="Could not read the uploaded file as an Excel workbook"
        )
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    field_by_col = {
        idx: COLUMN_MAP[h] for idx, h in enumerate(header) if h in COLUMN_MAP
    }

    missing_columns = {"name", "designation", "email", "phone", "whatsapp_number",
                        "place_of_work", "participant_type"} - set(field_by_col.values())
    if missing_columns:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(sorted(missing_columns))}",
        )

    source_value = "website" if source_type == "cme_website" else "import"
    data_rows = rows[1:]

    # The whole batch (event lookup through every row's writes) runs inside one
    # transaction so a concurrent import/registration can't race past these
    # per-row duplicate checks (same class of bug fixed in registrations/
    # attendance/certificates — see excel_store.transaction()).
    with store.transaction() as tx:
        event = tx.find_row("Events", "event_id", event_id)
        if event is None:
            raise HTTPException(status_code=404, detail="Event not found")

        participants_by_email = {
            p["email"].lower(): p for p in tx.list_rows("Participants") if p.get("email")
        }
        existing_registrations = {
            (r["participant_id"], r["event_id"]) for r in tx.list_rows("Registrations")
        }

        errors: list[dict] = []

        for row_number, row in enumerate(data_rows, start=2):
            if row is None or all(cell in (None, "") for cell in row):
                continue

            data = {}
            for idx, field in field_by_col.items():
                value = row[idx] if idx < len(row) else None
                data[field] = str(value).strip() if value not in (None, "") else None

            row_errors = [
                f"Missing {field}" for field in REQUIRED_FIELDS if not data.get(field)
            ]

            participant_type = (data.get("participant_type") or "").strip().title()
            if participant_type not in ("Faculty", "Delegate"):
                row_errors.append("participant_type must be Faculty or Delegate")
            else:
                data["participant_type"] = participant_type

            if event["cme_credits"] and not data.get("medical_license_no"):
                row_errors.append(
                    "Medical license number is required for CME-credit events"
                )

            if row_errors:
                errors.append(
                    {"row_number": row_number, "error_message": "; ".join(row_errors)}
                )
                continue

            email_key = data["email"].lower()
            participant = participants_by_email.get(email_key)
            if participant is None:
                participant = {
                    "participant_id": store.new_id(),
                    "name": data["name"],
                    "designation": data["designation"],
                    "email": data["email"],
                    "phone": data["phone"],
                    "whatsapp_number": data["whatsapp_number"],
                    "place_of_work": data["place_of_work"],
                    "country": data.get("country"),
                    "medical_license_no": data.get("medical_license_no"),
                    "participant_type": data["participant_type"],
                    "source": source_value,
                    "created_at": store.now_iso(),
                }
                tx.append_row("Participants", participant)
                participants_by_email[email_key] = participant

            if (participant["participant_id"], event_id) in existing_registrations:
                errors.append(
                    {
                        "row_number": row_number,
                        "error_message": f"{data['email']} is already registered for this event",
                    }
                )
                continue

            registration = {
                "registration_id": store.new_id(),
                "participant_id": participant["participant_id"],
                "event_id": event_id,
                "source": source_value,
                "registered_at": store.now_iso(),
            }
            tx.append_row("Registrations", registration)
            existing_registrations.add((participant["participant_id"], event_id))

        batch_id = store.new_id()
        batch = {
            "batch_id": batch_id,
            "source_file": file.filename or "upload.xlsx",
            "source_type": source_type,
            "imported_at": store.now_iso(),
            "imported_by": current_user["email"],
            "row_count": len(data_rows),
            "error_count": len(errors),
        }
        tx.append_row("ImportBatches", batch)
        for err in errors:
            tx.append_row("ImportErrors", {"batch_id": batch_id, **err})

        return {"batch": batch, "errors": errors}


@router.get("/batches", response_model=list[ImportBatch])
def list_batches():
    return store.list_rows("ImportBatches")


@router.get("/batches/{batch_id}", response_model=ImportResult)
def get_batch(batch_id: str):
    batch = store.find_row("ImportBatches", "batch_id", batch_id)
    if batch is None:
        raise HTTPException(status_code=404, detail="Import batch not found")
    errors = [e for e in store.list_rows("ImportErrors") if e["batch_id"] == batch_id]
    return {"batch": batch, "errors": errors}

"""Data-access layer over the central Excel workbook (see docs/excel-schema.md).

All reads and writes go through this module so callers never touch the workbook
file directly. Concurrency: within this process a threading.Lock serializes
access; a FileLock additionally guards the workbook file itself, since a
production deployment may run more than one backend worker process.

Any check-then-act sequence (e.g. "reject if already registered, else
insert") MUST use transaction() to hold the lock across the whole sequence —
a single list_rows()/append_row() pair each acquire and release the lock
separately, which is only safe for one-shot operations. See the concurrency
test in this project's history: without transaction(), concurrent requests
raced past each other's duplicate checks (confirmed for certificate
sequential numbering, which reads-then-appends twice per call).
"""

import uuid
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from threading import Lock

import openpyxl
from filelock import FileLock

DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"
WORKBOOK_PATH = DATA_DIR / "central.xlsx"
LOCK_PATH = DATA_DIR / "central.xlsx.lock"

SHEETS: dict[str, list[str]] = {
    "Participants": [
        "participant_id", "name", "designation", "email", "phone",
        "whatsapp_number", "place_of_work", "country", "medical_license_no",
        "participant_type", "source", "created_at",
    ],
    "Events": [
        "event_id", "event_name", "event_date", "venue", "organizing_doctors",
        "department", "cme_credits", "approx_duration_hours",
    ],
    "Registrations": [
        "registration_id", "participant_id", "event_id", "source", "registered_at",
    ],
    "Attendance": [
        "attendance_id", "registration_id", "status", "sign_in_time",
        "sign_in_signature_ref", "sign_out_time", "sign_out_signature_ref",
        "device_id",
    ],
    "Certificates": [
        "certificate_id", "certificate_no", "event_id", "participant_id",
        "delivery_status", "issued_at",
    ],
    "ImportBatches": [
        "batch_id", "source_file", "source_type", "imported_at", "imported_by",
        "row_count", "error_count",
    ],
    "ImportErrors": ["batch_id", "row_number", "error_message"],
    "Users": ["user_id", "name", "role", "email", "password_hash"],
    "AuditLogs": [
        "log_id", "user_id", "action", "target_ref", "timestamp", "details",
    ],
}

_process_lock = Lock()


def new_id() -> str:
    return uuid.uuid4().hex


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def init_workbook() -> None:
    """Create the central workbook with all sheets/headers if it doesn't exist yet."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if WORKBOOK_PATH.exists():
        return
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, headers in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
    wb.save(WORKBOOK_PATH)


def _file_lock() -> FileLock:
    return FileLock(str(LOCK_PATH), timeout=10)


class Transaction:
    """A single load-mutate-save cycle over the workbook. Reads see the
    transaction's own uncommitted writes; nothing is persisted until the
    transaction exits cleanly (an exception raised inside the `with` block
    discards all writes made in it)."""

    def __init__(self, wb: openpyxl.Workbook):
        self._wb = wb
        self.dirty = False

    def list_rows(self, sheet_name: str) -> list[dict]:
        headers = SHEETS[sheet_name]
        ws = self._wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            rows.append(dict(zip(headers, row)))
        return rows

    def find_row(self, sheet_name: str, id_column: str, id_value: str) -> dict | None:
        for row in self.list_rows(sheet_name):
            if row.get(id_column) == id_value:
                return row
        return None

    def append_row(self, sheet_name: str, row: dict) -> dict:
        headers = SHEETS[sheet_name]
        ws = self._wb[sheet_name]
        ws.append([row.get(h) for h in headers])
        self.dirty = True
        return row

    def update_row(
        self, sheet_name: str, id_column: str, id_value: str, updates: dict
    ) -> dict | None:
        headers = SHEETS[sheet_name]
        id_idx = headers.index(id_column)
        ws = self._wb[sheet_name]
        for row_cells in ws.iter_rows(min_row=2):
            if row_cells[id_idx].value == id_value:
                for key, value in updates.items():
                    row_cells[headers.index(key)].value = value
                self.dirty = True
                return dict(zip(headers, (c.value for c in row_cells)))
        return None


@contextmanager
def transaction():
    """Hold the workbook lock across an entire read-check-write sequence.
    Commits (saves) only if the block completes without raising."""
    with _file_lock(), _process_lock:
        wb = openpyxl.load_workbook(WORKBOOK_PATH)
        tx = Transaction(wb)
        yield tx
        if tx.dirty:
            wb.save(WORKBOOK_PATH)


def list_rows(sheet_name: str) -> list[dict]:
    """Return every row of a sheet as a list of dicts, in sheet order."""
    with transaction() as tx:
        return tx.list_rows(sheet_name)


def find_row(sheet_name: str, id_column: str, id_value: str) -> dict | None:
    """Return the first row where id_column == id_value, or None."""
    with transaction() as tx:
        return tx.find_row(sheet_name, id_column, id_value)


def append_row(sheet_name: str, row: dict) -> dict:
    """Append a row to a sheet. Any header missing from `row` is stored as None."""
    with transaction() as tx:
        return tx.append_row(sheet_name, row)


def update_row(
    sheet_name: str, id_column: str, id_value: str, updates: dict
) -> dict | None:
    """Update the first row matching id_column == id_value. Returns the updated
    row, or None if no row matched."""
    with transaction() as tx:
        return tx.update_row(sheet_name, id_column, id_value, updates)
